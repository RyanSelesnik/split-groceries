import re
import sys

try:
    import pdfplumber
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pdfplumber"])
    import pdfplumber

try:
    import openpyxl
    from openpyxl.styles import Font, numbers
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, numbers


def parse_sainsburys_receipt(pdf_path: str, output_path: str) -> list[dict]:
    """
    Parse a Sainsbury's receipt PDF and output an xlsx for splitting groceries.

    Args:
        pdf_path: Path to the Sainsbury's receipt PDF.
        output_path: Path to write the output xlsx.

    Returns:
        List of parsed item dicts.
    """
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"

    # Fix PDF encoding issues (e.g. missing apostrophes)
    text = re.sub(r"\(cid:\d+\)", "'", text)

    # Extract the groceries section
    groceries_match = re.search(r"Groceries\s*\(\d+\s*items?\)", text)
    if not groceries_match:
        raise ValueError("Could not find 'Groceries' section in the receipt.")

    groceries_text = text[groceries_match.end():]

    # Stop at "Order summary" or end of relevant content
    end_match = re.search(r"Order summary|Food Information|Need any help", groceries_text)
    if end_match:
        groceries_text = groceries_text[:end_match.start()]

    # Parse items: lines like "1 Item Name £1.50" or "2 Item Name £4.40"
    # Price is always £X.XX at the end of a line
    # Quantity is a number at the start
    # Item names can span multiple lines before the price
    lines = groceries_text.strip().split("\n")

    items = []
    buffer_qty = None
    buffer_name_parts = []

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Check if this line ends with a price
        price_match = re.search(r"£(\d+\.\d{2})\s*$", line)

        if price_match:
            price_str = price_match.group(1)
            before_price = line[:price_match.start()].strip()

            # Check if this line starts with a quantity
            qty_match = re.match(r"^(\d+)\s+", before_price)

            if qty_match and buffer_qty is None:
                # Single-line item: "1 Item Name £1.50"
                qty = int(qty_match.group(1))
                name = before_price[qty_match.end():].strip()
                items.append({
                    "Item Name": name,
                    "Quantity": qty,
                    "Payed": float(price_str),
                })
            elif buffer_qty is not None:
                # Multi-line item: this is the last line with the price
                buffer_name_parts.append(before_price)
                name = " ".join(buffer_name_parts).strip()
                items.append({
                    "Item Name": name,
                    "Quantity": buffer_qty,
                    "Payed": float(price_str),
                })
                buffer_qty = None
                buffer_name_parts = []
            elif qty_match:
                # New item starts on this line but had a buffered item
                qty = int(qty_match.group(1))
                name = before_price[qty_match.end():].strip()
                items.append({
                    "Item Name": name,
                    "Quantity": qty,
                    "Payed": float(price_str),
                })
        else:
            # No price on this line — either start of a multi-line item or continuation
            qty_match = re.match(r"^(\d+)\s+", line)
            if qty_match:
                # Start of a new multi-line item
                buffer_qty = int(qty_match.group(1))
                buffer_name_parts = [line[qty_match.end():].strip()]
            elif buffer_qty is not None:
                # Continuation of a multi-line item
                buffer_name_parts.append(line.strip())

    # Check for delivery cost
    delivery_match = re.search(r"Delivery cost\s+£(\d+\.\d{2})", text)
    if delivery_match:
        items.append({
            "Item Name": "Delivery",
            "Quantity": 1,
            "Payed": float(delivery_match.group(1)),
        })

    write_xlsx(items, output_path)
    return items


def parse_tesco_receipt(pdf_path: str, output_path: str) -> list[dict]:
    """
    Parse a Tesco receipt PDF and output an xlsx for splitting groceries.

    Args:
        pdf_path: Path to the Tesco receipt PDF.
        output_path: Path to write the output xlsx.

    Returns:
        List of parsed item dicts.
    """
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"

    # Find the start of items (after "Rest of your items")
    items_match = re.search(r"Rest of your items", text)
    if not items_match:
        raise ValueError("Could not find 'Rest of your items' section in the receipt.")

    items_text = text[items_match.end():]

    # Stop at end of items section
    end_match = re.search(r"Shop from this order|Order help|Rejected vouchers", items_text)
    if end_match:
        items_text = items_text[:end_match.start()]

    raw_lines = items_text.strip().split("\n")

    # Pre-filter noise lines (page headers/footers, clubcard, deal descriptions, section headers)
    section_headers = {"Fridge", "Freezer", "Cupboard", "Bakery", "Fruit & Veg",
                       "Drinks", "Food Cupboard"}
    noise = re.compile(
        r"https://www\.tesco\.com/"
        r"|\d{2}/\d{2}/\d{4},\s+\d{2}:\d{2}\s+My orders"
        r"|Clubcard Price"
        r"|^Any \d+"
    )
    lines = []
    for raw in raw_lines:
        stripped = raw.strip()
        if not stripped or stripped in section_headers or noise.search(stripped):
            continue
        lines.append(stripped)

    items = []
    i = 0
    while i < len(lines):
        line = lines[i]

        # Try to match: item name, then £price on next line, then Quantity: N
        price_match = None
        qty_match = None

        if i + 1 < len(lines):
            price_match = re.match(r"^£(\d+\.\d{2})$", lines[i + 1])
        if i + 2 < len(lines):
            qty_match = re.match(r"^Quantity:\s*(\d+)$", lines[i + 2])

        if price_match and qty_match:
            name = line
            price = float(price_match.group(1))
            qty = int(qty_match.group(1))
            items.append({
                "Item Name": name,
                "Quantity": qty,
                "Payed": price,
            })
            i += 3
        else:
            i += 1

    # Check for delivery cost ("Pick, pack and deliver £X.XX")
    delivery_match = re.search(r"Pick, pack and deliver\s+£(\d+\.\d{2})", text)
    if delivery_match:
        items.append({
            "Item Name": "Delivery",
            "Quantity": 1,
            "Payed": float(delivery_match.group(1)),
        })

    write_xlsx(items, output_path)
    return items


def write_xlsx(items: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense_Breakdown"

    headers = ["Item Name", "Quantity", "Payed", "Ryan", "Rael", "Gia",
               "Ryan owes", "Rael owes", "Gia owes"]
    bold = Font(bold=True)
    gbp_fmt = '£#,##0.00'

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold

    for i, item in enumerate(items):
        row = i + 2  # data starts at row 2
        ws.cell(row=row, column=1, value=item["Item Name"])
        ws.cell(row=row, column=2, value=item["Quantity"])
        ws.cell(row=row, column=3, value=item["Payed"]).number_format = gbp_fmt
        ws.cell(row=row, column=4, value=0)  # Ryan
        ws.cell(row=row, column=5, value=0)  # Rael
        ws.cell(row=row, column=6, value=0)  # Gia

        # Formulas: person_owes = IF(person>0, price*person/(ryan+rael+gia), 0)
        # D=Ryan, E=Rael, F=Gia, C=Payed
        ws.cell(row=row, column=7).value = f"=IF(D{row}>0,C{row}*D{row}/(D{row}+E{row}+F{row}),0)"
        ws.cell(row=row, column=7).number_format = gbp_fmt
        ws.cell(row=row, column=8).value = f"=IF(E{row}>0,C{row}*E{row}/(D{row}+E{row}+F{row}),0)"
        ws.cell(row=row, column=8).number_format = gbp_fmt
        ws.cell(row=row, column=9).value = f"=IF(F{row}>0,C{row}*F{row}/(D{row}+E{row}+F{row}),0)"
        ws.cell(row=row, column=9).number_format = gbp_fmt

    # TOTAL row
    total_row = len(items) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = bold
    ws.cell(row=total_row, column=3).value = f"=SUM(C2:C{total_row - 1})"
    ws.cell(row=total_row, column=3).number_format = gbp_fmt
    ws.cell(row=total_row, column=3).font = bold
    ws.cell(row=total_row, column=7).value = f"=SUM(G2:G{total_row - 1})"
    ws.cell(row=total_row, column=7).number_format = gbp_fmt
    ws.cell(row=total_row, column=7).font = bold
    ws.cell(row=total_row, column=8).value = f"=SUM(H2:H{total_row - 1})"
    ws.cell(row=total_row, column=8).number_format = gbp_fmt
    ws.cell(row=total_row, column=8).font = bold
    ws.cell(row=total_row, column=9).value = f"=SUM(I2:I{total_row - 1})"
    ws.cell(row=total_row, column=9).number_format = gbp_fmt
    ws.cell(row=total_row, column=9).font = bold

    # Auto-fit column A width
    ws.column_dimensions['A'].width = 60

    wb.save(output_path)
    total = sum(item["Payed"] for item in items)
    print(f"Parsed {len(items)} items, total: £{total:.2f}")
    print(f"xlsx written to: {output_path}")


def detect_store(pdf_path: str) -> str:
    with pdfplumber.open(pdf_path) as pdf:
        first_page = pdf.pages[0].extract_text() or ""
    if "Tesco" in first_page:
        return "tesco"
    if "Sainsbury" in first_page:
        return "sainsburys"
    raise ValueError("Could not detect store from receipt. Supported: Sainsbury's, Tesco.")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(f"Usage: python {sys.argv[0]} <receipt.pdf> <output.xlsx>")
        sys.exit(1)
    pdf_path = sys.argv[1]
    output_path = sys.argv[2]
    store = detect_store(pdf_path)
    if store == "tesco":
        parse_tesco_receipt(pdf_path, output_path)
    else:
        parse_sainsburys_receipt(pdf_path, output_path)
